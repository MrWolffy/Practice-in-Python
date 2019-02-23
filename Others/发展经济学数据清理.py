import openpyxl
import re

# 去世年龄为100
# 子女上幼儿园：受教育程度 = 2

title = []
it = [0, 0]


def nextIter():
    global it
    it[0] += 1
    it[1] += 1
    return it[0]


def to_int(value):
    try:
        tmpval = int(value)
    except ValueError:
        tmpval = str(value)
    except:
        tmpval = ""
    return tmpval


def to_float(value):
    try:
        tmpval = float(value)
    except ValueError:
        tmpval = str(value)
    except:
        tmpval = ""
    return tmpval


def makeTitle():
    global title
    title = ["", "q0", "q1"]
    for i in range(1, 11):
        title.append("q2_" + str(i))
    for i in range(1, 9):
        title.append("q3_" + str(i))
    for i in range(1, 41):
        title.append("q4_" + str(i))
    for i in range(1, 3):
        title.append("q5_" + str(i))
    for i in range(1, 5):
        title.append("q6_" + str(i))
    for i in range(1, 6):
        title.append("q7_" + str(i))
    title.append("q8_1")
    for i in range(1, 12):
        title.append("q8_2_" + str(i))
    for i in range(1, 3):
        title.append("q9_" + str(i))
    for i in range(1, 12):
        title.append("q9_3_" + str(i))
    for i in range(4, 7):
        title.append("q9_" + str(i))
    for i in range(1, 7):
        title.append("q9_7_" + str(i))
    title.append("q9_8")
    for i in range(1, 4):
        title.append("q10_" + str(i))
    for i in range(1, 4):
        title.append("q11_" + str(i))
    for i in range(12, 15):
        title.append("q" + str(i) + "")
    for i in range(1, 11):
        title.append("q15_1_" + str(i))
    title.append("q16")
    title.append("q17")
    for i in range(1, 8):
        title.append("q18_1_" + str(i))
    title.append("q19")
    for i in range(1, 7):
        title.append("q20_1_" + str(i))
    title.append("q21")
    for i in range(1, 9):
        title.append("q22_1_" + str(i))
    for i in range(1, 9):
        title.append("q22_2_" + str(i))
    title.append("q23")
    for i in range(1, 5):
        title.append("q24_" + str(i))
    for i in range(1, 3):
        title.append("q25_" + str(i))
    for i in range(1, 3):
        title.append("q26_" + str(i))
    for i in range(1, 7):
        title.append("q27_1_" + str(i))
    for i in range(1, 9):
        title.append("q28_1_" + str(i))
    for i in range(1, 8):
        title.append("q29_1_" + str(i))
    for i in range(1, 3):
        title.append("q30_" + str(i))
    for i in range(1, 7):
        title.append("q30_3_" + str(i))
    for i in range(1, 8):
        title.append("q31_1_" + str(i))
    for i in range(1, 8):
        title.append("q31_2_" + str(i))
    title.append("q32")


def makeReference(sheet):
    pass


def clearQ0(sheet, newsheet):
    newsheet["A1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["A" + str(i - 2)] = str(sheet.cell(row=i, column=it[1]).value)


def clearQ1(sheet, newsheet):
    newsheet["B1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["B" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)


def clearQ2(sheet, newsheet):
    newsheet["C1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["C" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["D1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = sheet.cell(row=i, column=it[1]).value
        if tmpval == "河北 ":
            tmpval = "河北"
        elif tmpval == "河南 " or tmpval == "河南0":
            tmpval = "河南"
        elif tmpval == "湖北 ":
            tmpval = "湖北"
        elif tmpval == "内蒙":
            tmpval = "内蒙古"
        newsheet["D" + str(i - 2)] = tmpval

    newsheet["E1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["E" + str(i - 2)] = sheet.cell(row=i, column=it[1]).value

    newsheet["F1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["F" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["G1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = -1
        newsheet["G" + str(i - 2)] = tmpval

    newsheet["H1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == int and tmpval > 4:
            tmpval = -1
        newsheet["H" + str(i - 2)] = tmpval

    newsheet["I1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval is not None:
            if tmpval == "﹣1" or (type(tmpval) == int and tmpval > 3):
                tmpval = -1
            if type(tmpval) == str and tmpval != "":
                if "2" in tmpval:
                    tmpval = 2
                elif "3" in tmpval:
                    tmpval = 3
        newsheet["I" + str(i - 2)] = tmpval

    newsheet["J1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval is not None:
            if type(tmpval) == int and tmpval == 185:
                tmpval = -1
        newsheet["J" + str(i - 2)] = tmpval

    newsheet["K1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval is not None:
            if type(tmpval) == int and tmpval < 100:
                tmpval = -1
        newsheet["K" + str(i - 2)] = tmpval

    newsheet["L1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "62,5":
            tmpval = 62.5
        newsheet["L" + str(i - 2)] = tmpval


def clearQ3(sheet, newsheet):
    newsheet["M1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval is not None:
            if tmpval == 6:
                tmpval = -1
        newsheet["M" + str(i - 2)] = tmpval

    newsheet["N1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "2000？还未出生":
            tmpval = -1
        newsheet["N" + str(i - 2)] = tmpval

    newsheet["O1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "2—3":
            tmpval = 2.5
        elif tmpval == "365`":
            tmpval = 365
        elif type(tmpval) is str:
            tmpval = None
        newsheet["O" + str(i - 2)] = tmpval

    newsheet["P1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "4.5分":
            tmpval = 0.45
        elif tmpval == "20多倾":
            tmpval = 2000
        elif tmpval == "40（注：小亩）":
            tmpval = 40
        elif tmpval == "45（注：小亩）":
            tmpval = 45
        newsheet["P" + str(i - 2)] = tmpval

    newsheet["Q1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "4.5分":
            tmpval = 0.45
        elif tmpval == "2000(草原)":
            tmpval = 2000
        newsheet["Q" + str(i - 2)] = tmpval

    newsheet["R1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["R" + str(i - 2)] = to_float(sheet.cell(row=i, column=it[1]).value)

    newsheet["S1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "2.6（弟弟）":
            tmpval = 2.6
        newsheet["S" + str(i - 2)] = tmpval

    newsheet["T1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["T" + str(i - 2)] = to_float(sheet.cell(row=i, column=it[1]).value)


def clearQ4(sheet, newsheet):
    newsheet["U1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["U" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["V1"] = title[nextIter()]
    for i in range(4, 1002):
        try:
            newsheet["V" + str(i - 2)] = int(sheet.cell(row=i, column=it[1]).value)
        except:
            newsheet["V" + str(i - 2)] = ""

    newsheet["W1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        newsheet["W" + str(i - 2)] = tmpval

    newsheet["X1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["X" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["Y1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        elif tmpval == "去世":
            tmpval = 100
        newsheet["Y" + str(i - 2)] = tmpval

    newsheet["Z1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        newsheet["Z" + str(i - 2)] = tmpval

    newsheet["AA1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        newsheet["AA" + str(i - 2)] = tmpval

    newsheet["AB1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        newsheet["AB" + str(i - 2)] = tmpval

    newsheet["AC1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str:
            if "去世" in tmpval or "过世" in tmpval:
                tmpval = 100
            else:
                tmpval = ""
        newsheet["AC" + str(i - 2)] = tmpval

    newsheet["AD1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "否":
            tmpval = 2
        newsheet["AD" + str(i - 2)] = tmpval

    newsheet["AE1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = 2
        newsheet["AE" + str(i - 2)] = tmpval

    newsheet["AF1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "1 ":
            tmpval = 1
        elif tmpval == 44:
            tmpval = 4
        newsheet["AF" + str(i - 2)] = tmpval

    newsheet["AG1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "去世":
            tmpval = 100
        elif tmpval == "50+" or tmpval == "60+":
            tmpval = ""
        newsheet["AG" + str(i - 2)] = tmpval

    newsheet["AH1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["AH" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["AI1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        newsheet["AI" + str(i - 2)] = tmpval

    newsheet["AJ1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["AJ" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["AK1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        elif type(tmpval) == str:
            if "已故" in tmpval or "去世" in tmpval or "亡故" in tmpval:
                tmpval = ""
            else:
                tmpval = 4
        newsheet["AK" + str(i - 2)] = tmpval

    newsheet["AL1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["AL" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["AM1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        newsheet["AM" + str(i - 2)] = tmpval

    newsheet["AN1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["AN" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["AO1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "男":
            tmpval = 1
        elif tmpval == "女":
            tmpval = 2
        newsheet["AO" + str(i - 2)] = tmpval

    newsheet["AP1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "13个月":
            tmpval = 13 / 12
        elif tmpval == "1岁8个月":
            tmpval = 1 + 8 / 12
        elif tmpval == "1周":
            tmpval = 1 / 52
        elif tmpval == "2周":
            tmpval = 2 / 52
        elif tmpval == "去世":
            tmpval = 100
        newsheet["AP" + str(i - 2)] = tmpval

    newsheet["AQ1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "幼儿园":
            tmpval = 2
        elif tmpval == "高中":
            tmpval = 5
        newsheet["AQ" + str(i - 2)] = tmpval

    newsheet["AR1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = 3
        newsheet["AR" + str(i - 2)] = tmpval

    newsheet["AS1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str:
            if "2" in tmpval:
                tmpval = 2
            elif "3" in tmpval:
                tmpval = 3
        newsheet["AS" + str(i - 2)] = tmpval

    newsheet["AT1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "男":
            tmpval = 1
        elif tmpval == "女":
            tmpval = 2
        elif tmpval == " ":
            tmpval = ""
        elif tmpval == "子女2去世":
            tmpval = 100
        newsheet["AT" + str(i - 2)] = tmpval

    newsheet["AU1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        elif tmpval == "2月半":
            tmpval = 2.5 / 12
        elif tmpval == "6个月":
            tmpval = 0.5
        newsheet["AU" + str(i - 2)] = tmpval

    newsheet["AV1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        elif tmpval == "3东莞":
            tmpval = 3
        elif tmpval == "幼儿园":
            tmpval = 2
        elif tmpval == "内蒙":
            tmpval = -1
        newsheet["AV" + str(i - 2)] = tmpval

    newsheet["AW1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        elif type(tmpval) == str and tmpval != "":
            tmpval = 3
        newsheet["AW" + str(i - 2)] = tmpval

    newsheet["AX1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        elif type(tmpval) == str and tmpval != "":
            tmpval = 3
        newsheet["AX" + str(i - 2)] = tmpval

    newsheet["AY1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "男":
            tmpval = 1
        elif tmpval == "女":
            tmpval = 2
        newsheet["AY" + str(i - 2)] = tmpval

    newsheet["AZ1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["AZ" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["BA1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["BA" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["BB1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = 3
        newsheet["BB" + str(i - 2)] = tmpval

    newsheet["BC1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str:
            if "2" in tmpval:
                tmpval = 2
            elif tmpval != "":
                tmpval = 3
        newsheet["BC" + str(i - 2)] = tmpval

    newsheet["BD1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "女":
            tmpval = 2
        newsheet["BD" + str(i - 2)] = tmpval

    newsheet["BE1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str:
            tmpval = ""
        newsheet["BE" + str(i - 2)] = tmpval

    newsheet["BF1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str:
            tmpval = ""
        newsheet["BF" + str(i - 2)] = tmpval

    newsheet["BG1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = 3
        newsheet["BG" + str(i - 2)] = tmpval

    newsheet["BH1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = 3
        newsheet["BH" + str(i - 2)] = tmpval


def clearQ5(sheet, newsheet):
    newsheet["BI1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["BI" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["BJ1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["BJ" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)


def clearQ6(sheet, newsheet):
    newsheet["BK1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["BK" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["BL1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["BL" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["BM1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "0 ":
            tmpval = 0
        newsheet["BM" + str(i - 2)] = tmpval

    newsheet["BN1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = -1
        newsheet["BN" + str(i - 2)] = tmpval


def clearQ7(sheet, newsheet):
    newsheet["BO1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["BO" + str(i - 2)] = to_float(sheet.cell(row=i, column=it[1]).value)

    newsheet["BP1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["BP" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["BQ1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = -1
        newsheet["BQ" + str(i - 2)] = tmpval

    newsheet["BR1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["BR" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["BS1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = -1
        newsheet["BS" + str(i - 2)] = tmpval


def clearQ8(sheet, newsheet):
    newsheet["BT1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = 8
        newsheet["BT" + str(i - 2)] = tmpval

    newsheet["BU1"] = title[nextIter()]
    newsheet["BV1"] = title[nextIter()]
    newsheet["BW1"] = title[nextIter()]
    newsheet["BX1"] = title[nextIter()]
    newsheet["BY1"] = title[nextIter()]
    newsheet["BZ1"] = title[nextIter()]
    newsheet["CA1"] = title[nextIter()]
    newsheet["CB1"] = title[nextIter()]
    newsheet["CC1"] = title[nextIter()]
    newsheet["CD1"] = title[nextIter()]
    newsheet["CE1"] = title[nextIter()]
    it[1] -= 10
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval == "-1":
            newsheet["BU" + str(i - 2)] = -1
            newsheet["BV" + str(i - 2)] = -1
            newsheet["BW" + str(i - 2)] = -1
            newsheet["BX" + str(i - 2)] = -1
            newsheet["BY" + str(i - 2)] = -1
            newsheet["BZ" + str(i - 2)] = -1
            newsheet["CA" + str(i - 2)] = -1
            newsheet["CB" + str(i - 2)] = -1
            newsheet["CC" + str(i - 2)] = -1
            newsheet["CD" + str(i - 2)] = -1
            newsheet["CE" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["BU" + str(i - 2)] = ""
            newsheet["BV" + str(i - 2)] = ""
            newsheet["BW" + str(i - 2)] = ""
            newsheet["BX" + str(i - 2)] = ""
            newsheet["BY" + str(i - 2)] = ""
            newsheet["BZ" + str(i - 2)] = ""
            newsheet["CA" + str(i - 2)] = ""
            newsheet["CB" + str(i - 2)] = ""
            newsheet["CC" + str(i - 2)] = ""
            newsheet["CD" + str(i - 2)] = ""
            newsheet["CE" + str(i - 2)] = ""
            continue
        newsheet["BU" + str(i - 2)] = int(re.match("1[^01]", tmpval) is not None)
        newsheet["BV" + str(i - 2)] = int("2" in tmpval)
        newsheet["BW" + str(i - 2)] = int("3" in tmpval)
        newsheet["BX" + str(i - 2)] = int("4" in tmpval)
        newsheet["BY" + str(i - 2)] = int("5" in tmpval)
        newsheet["BZ" + str(i - 2)] = int("6" in tmpval)
        newsheet["CA" + str(i - 2)] = int("7" in tmpval)
        newsheet["CB" + str(i - 2)] = int("8" in tmpval)
        newsheet["CC" + str(i - 2)] = int("9" in tmpval)
        newsheet["CD" + str(i - 2)] = int("10" in tmpval)
        newsheet["CE" + str(i - 2)] = int("11" in tmpval)


def clearQ9(sheet, newsheet):
    newsheet["CF1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str:
            tmpval = -1
        newsheet["CF" + str(i - 2)] = tmpval

    newsheet["CG1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "100000（首付，总价值600000）":
            tmpval = 600000
        elif tmpval == "100000左右":
            tmpval = 100000
        elif tmpval == "1-20000":
            tmpval = 15000
        elif tmpval == "4-50000":
            tmpval = 45000
        elif tmpval == "7-80000":
            tmpval = 75000
        elif type(tmpval) == str and tmpval != "":
            tmpval = -1
        newsheet["CG" + str(i - 2)] = tmpval

    newsheet["CH1"] = title[nextIter()]
    newsheet["CI1"] = title[nextIter()]
    newsheet["CJ1"] = title[nextIter()]
    newsheet["CK1"] = title[nextIter()]
    newsheet["CL1"] = title[nextIter()]
    newsheet["CM1"] = title[nextIter()]
    newsheet["CN1"] = title[nextIter()]
    newsheet["CO1"] = title[nextIter()]
    newsheet["CP1"] = title[nextIter()]
    newsheet["CQ1"] = title[nextIter()]
    newsheet["CR1"] = title[nextIter()]
    it[1] -= 10
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval == "-1":
            newsheet["CH" + str(i - 2)] = -1
            newsheet["CI" + str(i - 2)] = -1
            newsheet["CJ" + str(i - 2)] = -1
            newsheet["CK" + str(i - 2)] = -1
            newsheet["CL" + str(i - 2)] = -1
            newsheet["CM" + str(i - 2)] = -1
            newsheet["CN" + str(i - 2)] = -1
            newsheet["CO" + str(i - 2)] = -1
            newsheet["CP" + str(i - 2)] = -1
            newsheet["CQ" + str(i - 2)] = -1
            newsheet["CR" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["CH" + str(i - 2)] = ""
            newsheet["CI" + str(i - 2)] = ""
            newsheet["CJ" + str(i - 2)] = ""
            newsheet["CK" + str(i - 2)] = ""
            newsheet["CL" + str(i - 2)] = ""
            newsheet["CM" + str(i - 2)] = ""
            newsheet["CN" + str(i - 2)] = ""
            newsheet["CO" + str(i - 2)] = ""
            newsheet["CP" + str(i - 2)] = ""
            newsheet["CQ" + str(i - 2)] = ""
            newsheet["CR" + str(i - 2)] = ""
            continue
        newsheet["CH" + str(i - 2)] = int(re.match("1[^01]", tmpval) is not None)
        newsheet["CI" + str(i - 2)] = int("2" in tmpval)
        newsheet["CJ" + str(i - 2)] = int("3" in tmpval)
        newsheet["CK" + str(i - 2)] = int("4" in tmpval)
        newsheet["CL" + str(i - 2)] = int("5" in tmpval)
        newsheet["CM" + str(i - 2)] = int("6" in tmpval)
        newsheet["CN" + str(i - 2)] = int("7" in tmpval)
        newsheet["CO" + str(i - 2)] = int("8" in tmpval)
        newsheet["CP" + str(i - 2)] = int("9" in tmpval)
        newsheet["CQ" + str(i - 2)] = int("10" in tmpval)
        newsheet["CR" + str(i - 2)] = int("11" in tmpval)

    newsheet["CS1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["CS" + str(i - 2)] = to_float(sheet.cell(row=i, column=it[1]).value)

    newsheet["CT1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = -1
        newsheet["CT" + str(i - 2)] = tmpval

    newsheet["CU1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["CU" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["CV1"] = title[nextIter()]
    newsheet["CW1"] = title[nextIter()]
    newsheet["CX1"] = title[nextIter()]
    newsheet["CY1"] = title[nextIter()]
    newsheet["CZ1"] = title[nextIter()]
    newsheet["DA1"] = title[nextIter()]
    it[1] -= 5
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval == "-1":
            newsheet["CV" + str(i - 2)] = -1
            newsheet["CW" + str(i - 2)] = -1
            newsheet["CX" + str(i - 2)] = -1
            newsheet["CY" + str(i - 2)] = -1
            newsheet["CZ" + str(i - 2)] = -1
            newsheet["DA" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["CV" + str(i - 2)] = ""
            newsheet["CW" + str(i - 2)] = ""
            newsheet["CX" + str(i - 2)] = ""
            newsheet["CY" + str(i - 2)] = ""
            newsheet["CZ" + str(i - 2)] = ""
            newsheet["DA" + str(i - 2)] = ""
            continue
        newsheet["CV" + str(i - 2)] = int(re.match("^1", tmpval) is not None)
        newsheet["CW" + str(i - 2)] = int("2" in tmpval)
        newsheet["CX" + str(i - 2)] = int("3" in tmpval)
        newsheet["CY" + str(i - 2)] = int("4" in tmpval)
        newsheet["CZ" + str(i - 2)] = int("5" in tmpval)
        newsheet["DA" + str(i - 2)] = int("6" in tmpval or tmpval == "因为长时间生活在北京" or tmpval == "丈夫的房子")

    newsheet["DB1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["DB" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)


def clearQ10(sheet, newsheet):
    newsheet["DC1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["DC" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["DD1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["DD" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["DE1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "10（2018回家干农活）":
            tmpval = 10
        elif tmpval == "3~4":
            tmpval = 3.5
        newsheet["DE" + str(i - 2)] = tmpval


def clearQ11(sheet, newsheet):
    newsheet["DF1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_float(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "7~8":
            tmpval = 7
        newsheet["DF" + str(i - 2)] = tmpval

    newsheet["DG1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == 2107:
            tmpval = 2017
        elif tmpval == 20009:
            tmpval = 2009
        elif type(tmpval) == str and tmpval != "":
            tmpval = -1
        newsheet["DG" + str(i - 2)] = tmpval

    newsheet["DH1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["DH" + str(i - 2)] = str(sheet.cell(row=i, column=it[1]).value)


def clearQ12(sheet, newsheet):
    newsheet["DI1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = 9
        newsheet["DI" + str(i - 2)] = tmpval


def clearQ13(sheet, newsheet):
    newsheet["DJ1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = -1
        newsheet["DJ" + str(i - 2)] = tmpval


def clearQ14(sheet, newsheet):
    newsheet["DK1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = -1
        newsheet["DK" + str(i - 2)] = tmpval


def clearQ15(sheet, newsheet):
    newsheet["DL1"] = title[nextIter()]
    newsheet["DM1"] = title[nextIter()]
    newsheet["DN1"] = title[nextIter()]
    newsheet["DO1"] = title[nextIter()]
    newsheet["DP1"] = title[nextIter()]
    newsheet["DQ1"] = title[nextIter()]
    newsheet["DR1"] = title[nextIter()]
    newsheet["DS1"] = title[nextIter()]
    newsheet["DT1"] = title[nextIter()]
    newsheet["DU1"] = title[nextIter()]
    it[1] -= 9
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval in ["-1", "20000", "30000"]:
            newsheet["DL" + str(i - 2)] = -1
            newsheet["DM" + str(i - 2)] = -1
            newsheet["DN" + str(i - 2)] = -1
            newsheet["DO" + str(i - 2)] = -1
            newsheet["DP" + str(i - 2)] = -1
            newsheet["DQ" + str(i - 2)] = -1
            newsheet["DR" + str(i - 2)] = -1
            newsheet["DS" + str(i - 2)] = -1
            newsheet["DT" + str(i - 2)] = -1
            newsheet["DU" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["DL" + str(i - 2)] = ""
            newsheet["DM" + str(i - 2)] = ""
            newsheet["DN" + str(i - 2)] = ""
            newsheet["DO" + str(i - 2)] = ""
            newsheet["DP" + str(i - 2)] = ""
            newsheet["DQ" + str(i - 2)] = ""
            newsheet["DR" + str(i - 2)] = ""
            newsheet["DS" + str(i - 2)] = ""
            newsheet["DT" + str(i - 2)] = ""
            newsheet["DU" + str(i - 2)] = ""
            continue
        newsheet["DL" + str(i - 2)] = int(re.match("^1", tmpval) is not None)
        newsheet["DM" + str(i - 2)] = int("2" in tmpval)
        newsheet["DN" + str(i - 2)] = int("3" in tmpval)
        newsheet["DO" + str(i - 2)] = int("4" in tmpval)
        newsheet["DP" + str(i - 2)] = int("5" in tmpval)
        newsheet["DQ" + str(i - 2)] = int("6" in tmpval)
        newsheet["DR" + str(i - 2)] = int("7" in tmpval)
        newsheet["DS" + str(i - 2)] = int("8" in tmpval)
        newsheet["DT" + str(i - 2)] = int("9" in tmpval)
        newsheet["DU" + str(i - 2)] = int("10" in tmpval)


def clearQ16(sheet, newsheet):
    newsheet["DP1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "0刚开始工作没有工资":
            tmpval = 0
        elif tmpval == "100000（平均10万，今年赔了7万）":
            tmpval = 100000
        elif tmpval == "2-30000":
            tmpval = 25000
        newsheet["DP" + str(i - 2)] = tmpval


def clearQ17(sheet, newsheet):
    newsheet["DQ1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval not in [-1, 1, 2, 3]:
            tmpval = -1
        newsheet["DQ" + str(i - 2)] = tmpval


def clearQ18(sheet, newsheet):
    newsheet["DR1"] = title[nextIter()]
    newsheet["DS1"] = title[nextIter()]
    newsheet["DT1"] = title[nextIter()]
    newsheet["DU1"] = title[nextIter()]
    newsheet["DV1"] = title[nextIter()]
    newsheet["DW1"] = title[nextIter()]
    newsheet["DX1"] = title[nextIter()]
    it[1] -= 6
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval == "-1":
            newsheet["DR" + str(i - 2)] = -1
            newsheet["DS" + str(i - 2)] = -1
            newsheet["DT" + str(i - 2)] = -1
            newsheet["DU" + str(i - 2)] = -1
            newsheet["DV" + str(i - 2)] = -1
            newsheet["DW" + str(i - 2)] = -1
            newsheet["DX" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["DR" + str(i - 2)] = ""
            newsheet["DS" + str(i - 2)] = ""
            newsheet["DT" + str(i - 2)] = ""
            newsheet["DU" + str(i - 2)] = ""
            newsheet["DV" + str(i - 2)] = ""
            newsheet["DW" + str(i - 2)] = ""
            newsheet["DX" + str(i - 2)] = ""
            continue
        newsheet["DR" + str(i - 2)] = int(re.match("^1", tmpval) is not None)
        newsheet["DS" + str(i - 2)] = int("2" in tmpval)
        newsheet["DT" + str(i - 2)] = int("3" in tmpval)
        newsheet["DU" + str(i - 2)] = int("4" in tmpval)
        newsheet["DV" + str(i - 2)] = int("5" in tmpval)
        newsheet["DW" + str(i - 2)] = int("6" in tmpval)
        newsheet["DX" + str(i - 2)] = int("7" in tmpval)


def clearQ19(sheet, newsheet):
    newsheet["DY1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval not in [-1, 1, 2]:
            tmpval = -1
        newsheet["DY" + str(i - 2)] = tmpval


def clearQ20(sheet, newsheet):
    newsheet["DZ1"] = title[nextIter()]
    newsheet["EA1"] = title[nextIter()]
    newsheet["EB1"] = title[nextIter()]
    newsheet["EC1"] = title[nextIter()]
    newsheet["ED1"] = title[nextIter()]
    newsheet["EE1"] = title[nextIter()]
    it[1] -= 5
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval in ["-1", "10"]:
            newsheet["DZ" + str(i - 2)] = -1
            newsheet["EA" + str(i - 2)] = -1
            newsheet["EB" + str(i - 2)] = -1
            newsheet["EC" + str(i - 2)] = -1
            newsheet["ED" + str(i - 2)] = -1
            newsheet["EE" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["DZ" + str(i - 2)] = ""
            newsheet["EA" + str(i - 2)] = ""
            newsheet["EB" + str(i - 2)] = ""
            newsheet["EC" + str(i - 2)] = ""
            newsheet["ED" + str(i - 2)] = ""
            newsheet["EE" + str(i - 2)] = ""
            continue
        newsheet["DZ" + str(i - 2)] = int("1" in tmpval)
        newsheet["EA" + str(i - 2)] = int("2" in tmpval)
        newsheet["EB" + str(i - 2)] = int("3" in tmpval)
        newsheet["EC" + str(i - 2)] = int("4" in tmpval)
        newsheet["ED" + str(i - 2)] = int("5" in tmpval)
        newsheet["EE" + str(i - 2)] = int("6" in tmpval)


def clearQ21(sheet, newsheet):
    newsheet["EF1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == str and tmpval != "":
            tmpval = -1
        newsheet["EF" + str(i - 2)] = tmpval


def clearQ22(sheet, newsheet):
    newsheet["EG1"] = title[nextIter()]
    newsheet["EH1"] = title[nextIter()]
    newsheet["EI1"] = title[nextIter()]
    newsheet["EJ1"] = title[nextIter()]
    newsheet["EK1"] = title[nextIter()]
    newsheet["EL1"] = title[nextIter()]
    newsheet["EM1"] = title[nextIter()]
    newsheet["EN1"] = title[nextIter()]
    it[1] -= 7
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval in ["-1", "2.5", "10"]:
            newsheet["EG" + str(i - 2)] = -1
            newsheet["EH" + str(i - 2)] = -1
            newsheet["EI" + str(i - 2)] = -1
            newsheet["EJ" + str(i - 2)] = -1
            newsheet["EK" + str(i - 2)] = -1
            newsheet["EL" + str(i - 2)] = -1
            newsheet["EM" + str(i - 2)] = -1
            newsheet["EN" + str(i - 2)] = -1
            continue
        elif tmpval in ["", " ", "21题勾选的是老家县城，但是22选的是“是老家农村”对应的选项"]:
            newsheet["EG" + str(i - 2)] = ""
            newsheet["EH" + str(i - 2)] = ""
            newsheet["EI" + str(i - 2)] = ""
            newsheet["EJ" + str(i - 2)] = ""
            newsheet["EK" + str(i - 2)] = ""
            newsheet["EL" + str(i - 2)] = ""
            newsheet["EM" + str(i - 2)] = ""
            newsheet["EN" + str(i - 2)] = ""
            continue
        newsheet["EG" + str(i - 2)] = int(re.match("^1", tmpval) is not None)
        newsheet["EH" + str(i - 2)] = int("2" in tmpval)
        newsheet["EI" + str(i - 2)] = int("3" in tmpval)
        newsheet["EJ" + str(i - 2)] = int("4" in tmpval)
        newsheet["EK" + str(i - 2)] = int("5" in tmpval)
        newsheet["EL" + str(i - 2)] = int("6" in tmpval)
        newsheet["EM" + str(i - 2)] = int("7" in tmpval)
        newsheet["EN" + str(i - 2)] = int("8" in tmpval)

    newsheet["EO1"] = title[nextIter()]
    newsheet["EP1"] = title[nextIter()]
    newsheet["EQ1"] = title[nextIter()]
    newsheet["ER1"] = title[nextIter()]
    newsheet["ES1"] = title[nextIter()]
    newsheet["ET1"] = title[nextIter()]
    newsheet["EU1"] = title[nextIter()]
    newsheet["EV1"] = title[nextIter()]
    it[1] -= 7
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval in ["-1", "10"]:
            newsheet["EO" + str(i - 2)] = -1
            newsheet["EP" + str(i - 2)] = -1
            newsheet["EQ" + str(i - 2)] = -1
            newsheet["ER" + str(i - 2)] = -1
            newsheet["ES" + str(i - 2)] = -1
            newsheet["ET" + str(i - 2)] = -1
            newsheet["EU" + str(i - 2)] = -1
            newsheet["EV" + str(i - 2)] = -1
            continue
        elif tmpval in ["", " "]:
            newsheet["EO" + str(i - 2)] = ""
            newsheet["EP" + str(i - 2)] = ""
            newsheet["EQ" + str(i - 2)] = ""
            newsheet["ER" + str(i - 2)] = ""
            newsheet["ES" + str(i - 2)] = ""
            newsheet["ET" + str(i - 2)] = ""
            newsheet["EU" + str(i - 2)] = ""
            newsheet["EV" + str(i - 2)] = ""
            continue
        newsheet["EO" + str(i - 2)] = int(re.match("^1", tmpval) is not None)
        newsheet["EP" + str(i - 2)] = int("2" in tmpval)
        newsheet["EQ" + str(i - 2)] = int("3" in tmpval)
        newsheet["ER" + str(i - 2)] = int("4" in tmpval)
        newsheet["ES" + str(i - 2)] = int("5" in tmpval)
        newsheet["ET" + str(i - 2)] = int("6" in tmpval)
        newsheet["EU" + str(i - 2)] = int("7" in tmpval)
        newsheet["EV" + str(i - 2)] = int("8" in tmpval)


def clearQ23(sheet, newsheet):
    newsheet["EW1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval not in [-1, 1, 2, 3, 4, 5, 6, 7, 8]:
            tmpval = -1
        newsheet["EW" + str(i - 2)] = tmpval


def clearQ24(sheet, newsheet):
    newsheet["EX1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "1,":
            tmpval = 1
        newsheet["EX" + str(i - 2)] = tmpval

    newsheet["EY1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["EY" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["EZ1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if type(tmpval) == int and tmpval > 7:
            tmpval = -1
        newsheet["EZ" + str(i - 2)] = tmpval

    newsheet["FA1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "﹣1":
            tmpval = -1
        elif tmpval == "20W":
            tmpval = 200000
        elif tmpval == "25W":
            tmpval = 250000
        elif type(tmpval) == str:
            tmpval = ""
        newsheet["FA" + str(i - 2)] = tmpval


def clearQ25(sheet, newsheet):
    newsheet["FB1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["FB" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["FC1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == " ":
            tmpval = ""
        newsheet["FC" + str(i - 2)] = tmpval


def clearQ26(sheet, newsheet):
    newsheet["FD1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["FD" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["FE1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["FE" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)


def clearQ27(sheet, newsheet):
    newsheet["FF1"] = title[nextIter()]
    newsheet["FG1"] = title[nextIter()]
    newsheet["FH1"] = title[nextIter()]
    newsheet["FI1"] = title[nextIter()]
    newsheet["FJ1"] = title[nextIter()]
    newsheet["FK1"] = title[nextIter()]
    it[1] -= 5
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval == "-1":
            newsheet["FF" + str(i - 2)] = -1
            newsheet["FG" + str(i - 2)] = -1
            newsheet["FH" + str(i - 2)] = -1
            newsheet["FI" + str(i - 2)] = -1
            newsheet["FJ" + str(i - 2)] = -1
            newsheet["FK" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["FF" + str(i - 2)] = ""
            newsheet["FG" + str(i - 2)] = ""
            newsheet["FH" + str(i - 2)] = ""
            newsheet["FI" + str(i - 2)] = ""
            newsheet["FJ" + str(i - 2)] = ""
            newsheet["FK" + str(i - 2)] = ""
            continue
        newsheet["FF" + str(i - 2)] = int("1" in tmpval)
        newsheet["FG" + str(i - 2)] = int("2" in tmpval)
        newsheet["FH" + str(i - 2)] = int("3" in tmpval)
        newsheet["FI" + str(i - 2)] = int("4" in tmpval)
        newsheet["FJ" + str(i - 2)] = int("5" in tmpval)
        newsheet["FK" + str(i - 2)] = int("6" in tmpval)


def clearQ28(sheet, newsheet):
    newsheet["FL1"] = title[nextIter()]
    newsheet["FM1"] = title[nextIter()]
    newsheet["FN1"] = title[nextIter()]
    newsheet["FO1"] = title[nextIter()]
    newsheet["FP1"] = title[nextIter()]
    newsheet["FQ1"] = title[nextIter()]
    newsheet["FR1"] = title[nextIter()]
    newsheet["FS1"] = title[nextIter()]
    it[1] -= 7
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval == "-1":
            newsheet["FL" + str(i - 2)] = -1
            newsheet["FM" + str(i - 2)] = -1
            newsheet["FN" + str(i - 2)] = -1
            newsheet["FO" + str(i - 2)] = -1
            newsheet["FP" + str(i - 2)] = -1
            newsheet["FQ" + str(i - 2)] = -1
            newsheet["FR" + str(i - 2)] = -1
            newsheet["FS" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["FL" + str(i - 2)] = ""
            newsheet["FM" + str(i - 2)] = ""
            newsheet["FN" + str(i - 2)] = ""
            newsheet["FO" + str(i - 2)] = ""
            newsheet["FP" + str(i - 2)] = ""
            newsheet["FQ" + str(i - 2)] = ""
            newsheet["FR" + str(i - 2)] = ""
            newsheet["FS" + str(i - 2)] = ""
            continue
        newsheet["FL" + str(i - 2)] = int("1" in tmpval)
        newsheet["FM" + str(i - 2)] = int("2" in tmpval)
        newsheet["FN" + str(i - 2)] = int("3" in tmpval)
        newsheet["FO" + str(i - 2)] = int("4" in tmpval)
        newsheet["FP" + str(i - 2)] = int("5" in tmpval)
        newsheet["FQ" + str(i - 2)] = int("6" in tmpval)
        newsheet["FR" + str(i - 2)] = int("7" in tmpval)
        newsheet["FS" + str(i - 2)] = int("8" in tmpval)


def clearQ29(sheet, newsheet):
    newsheet["FT1"] = title[nextIter()]
    newsheet["FU1"] = title[nextIter()]
    newsheet["FV1"] = title[nextIter()]
    newsheet["FW1"] = title[nextIter()]
    newsheet["FX1"] = title[nextIter()]
    newsheet["FY1"] = title[nextIter()]
    newsheet["FZ1"] = title[nextIter()]
    it[1] -= 6
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval == "-1" or tmpval == "2018/7/1":
            newsheet["FT" + str(i - 2)] = -1
            newsheet["FU" + str(i - 2)] = -1
            newsheet["FV" + str(i - 2)] = -1
            newsheet["FW" + str(i - 2)] = -1
            newsheet["FX" + str(i - 2)] = -1
            newsheet["FY" + str(i - 2)] = -1
            newsheet["FZ" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["FT" + str(i - 2)] = ""
            newsheet["FU" + str(i - 2)] = ""
            newsheet["FV" + str(i - 2)] = ""
            newsheet["FW" + str(i - 2)] = ""
            newsheet["FX" + str(i - 2)] = ""
            newsheet["FY" + str(i - 2)] = ""
            newsheet["FZ" + str(i - 2)] = ""
            continue
        newsheet["FT" + str(i - 2)] = int("1" in tmpval)
        newsheet["FU" + str(i - 2)] = int("2" in tmpval)
        newsheet["FV" + str(i - 2)] = int("3" in tmpval)
        newsheet["FW" + str(i - 2)] = int("4" in tmpval)
        newsheet["FX" + str(i - 2)] = int("5" in tmpval)
        newsheet["FY" + str(i - 2)] = int("6" in tmpval)
        newsheet["FZ" + str(i - 2)] = int("7" in tmpval)


def clearQ30(sheet, newsheet):
    newsheet["GA1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == 6:
            tmpval = -1
        newsheet["GA" + str(i - 2)] = tmpval

    newsheet["GB1"] = title[nextIter()]
    for i in range(4, 1002):
        newsheet["GB" + str(i - 2)] = to_int(sheet.cell(row=i, column=it[1]).value)

    newsheet["GC1"] = title[nextIter()]
    newsheet["GD1"] = title[nextIter()]
    newsheet["GE1"] = title[nextIter()]
    newsheet["GF1"] = title[nextIter()]
    newsheet["GG1"] = title[nextIter()]
    newsheet["GH1"] = title[nextIter()]
    it[1] -= 5
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval == "-1":
            newsheet["GC" + str(i - 2)] = -1
            newsheet["GD" + str(i - 2)] = -1
            newsheet["GE" + str(i - 2)] = -1
            newsheet["GF" + str(i - 2)] = -1
            newsheet["GG" + str(i - 2)] = -1
            newsheet["GH" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["GC" + str(i - 2)] = ""
            newsheet["GD" + str(i - 2)] = ""
            newsheet["GE" + str(i - 2)] = ""
            newsheet["GF" + str(i - 2)] = ""
            newsheet["GG" + str(i - 2)] = ""
            newsheet["GH" + str(i - 2)] = ""
            continue
        newsheet["GC" + str(i - 2)] = int("1" in tmpval)
        newsheet["GD" + str(i - 2)] = int("2" in tmpval)
        newsheet["GE" + str(i - 2)] = int("3" in tmpval)
        newsheet["GF" + str(i - 2)] = int("4" in tmpval)
        newsheet["GG" + str(i - 2)] = int("5" in tmpval)
        newsheet["GH" + str(i - 2)] = int("6" in tmpval)


def clearQ31(sheet, newsheet):
    newsheet["GI1"] = title[nextIter()]
    newsheet["GJ1"] = title[nextIter()]
    newsheet["GK1"] = title[nextIter()]
    newsheet["GL1"] = title[nextIter()]
    newsheet["GM1"] = title[nextIter()]
    newsheet["GN1"] = title[nextIter()]
    newsheet["GO1"] = title[nextIter()]
    it[1] -= 6
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval == "-1":
            newsheet["GI" + str(i - 2)] = -1
            newsheet["GJ" + str(i - 2)] = -1
            newsheet["GK" + str(i - 2)] = -1
            newsheet["GL" + str(i - 2)] = -1
            newsheet["GM" + str(i - 2)] = -1
            newsheet["GN" + str(i - 2)] = -1
            newsheet["GO" + str(i - 2)] = -1
            continue
        elif tmpval == "":
            newsheet["GI" + str(i - 2)] = ""
            newsheet["GJ" + str(i - 2)] = ""
            newsheet["GK" + str(i - 2)] = ""
            newsheet["GL" + str(i - 2)] = ""
            newsheet["GM" + str(i - 2)] = ""
            newsheet["GN" + str(i - 2)] = ""
            newsheet["GO" + str(i - 2)] = ""
            continue
        newsheet["GI" + str(i - 2)] = int("1" in tmpval)
        newsheet["GJ" + str(i - 2)] = int("2" in tmpval)
        newsheet["GK" + str(i - 2)] = int("3" in tmpval)
        newsheet["GL" + str(i - 2)] = int("4" in tmpval)
        newsheet["GM" + str(i - 2)] = int("5" in tmpval)
        newsheet["GN" + str(i - 2)] = int("6" in tmpval)
        newsheet["GO" + str(i - 2)] = int("7" in tmpval)

    newsheet["GP1"] = title[nextIter()]
    newsheet["GQ1"] = title[nextIter()]
    newsheet["GR1"] = title[nextIter()]
    newsheet["GS1"] = title[nextIter()]
    newsheet["GT1"] = title[nextIter()]
    newsheet["GU1"] = title[nextIter()]
    newsheet["GV1"] = title[nextIter()]
    it[1] -= 6
    for i in range(4, 1002):
        try:
            tmpval = str(sheet.cell(row=i, column=it[1]).value)
        except:
            tmpval = ""
        if tmpval == "-1":
            newsheet["GP" + str(i - 2)] = -1
            newsheet["GQ" + str(i - 2)] = -1
            newsheet["GR" + str(i - 2)] = -1
            newsheet["GS" + str(i - 2)] = -1
            newsheet["GT" + str(i - 2)] = -1
            newsheet["GU" + str(i - 2)] = -1
            newsheet["GV" + str(i - 2)] = -1
            continue
        elif tmpval == "" or tmpval == " ":
            newsheet["GP" + str(i - 2)] = ""
            newsheet["GQ" + str(i - 2)] = ""
            newsheet["GR" + str(i - 2)] = ""
            newsheet["GS" + str(i - 2)] = ""
            newsheet["GT" + str(i - 2)] = ""
            newsheet["GU" + str(i - 2)] = ""
            newsheet["GV" + str(i - 2)] = ""
            continue
        newsheet["GP" + str(i - 2)] = int("1" in tmpval)
        newsheet["GQ" + str(i - 2)] = int("2" in tmpval)
        newsheet["GR" + str(i - 2)] = int("3" in tmpval)
        newsheet["GS" + str(i - 2)] = int("4" in tmpval)
        newsheet["GT" + str(i - 2)] = int("5" in tmpval)
        newsheet["GU" + str(i - 2)] = int("6" in tmpval)
        newsheet["GV" + str(i - 2)] = int("7" in tmpval)


def clearQ32(sheet, newsheet):
    newsheet["GW1"] = title[nextIter()]
    for i in range(4, 1002):
        tmpval = to_int(sheet.cell(row=i, column=it[1]).value)
        if tmpval == "1（前几年）":
            tmpval = 1
        newsheet["GW" + str(i - 2)] = tmpval




def main():
    workbook = openpyxl.load_workbook("数据汇总20181105.xlsx")
    sheet = workbook["Sheet1"]
    newwb = openpyxl.Workbook()
    newsheet = newwb["Sheet"]
    makeTitle()
    makeReference(sheet)
    for i in range(0, 33):
        eval("clearQ" + str(i) + "(sheet, newsheet)")
    newwb.save(r"data.xlsx")


if __name__ == '__main__':
    main()
